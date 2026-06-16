from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


FIELD_ALIASES = {
    "utmcontent": "utm_content",
    "utmsource": "utm_source",
    "utmcampaign": "utm_campaign",
    "utm_campaign": "utm_campaign",
    "utm_content": "utm_content",
    "utm_source": "utm_source",
    "formtitle": "form_title",
    "modulesource": "module_source",
}

SOURCE_FALLBACK_RULES = {
    "sms": ("MA", "sms"),
    "sms_5g": ("MA", "sms_5g"),
    "sms_5g_xcx": ("MA", "sms_5g"),
    "edm": ("MA", "edm"),
    "wechat_template": ("MA", "wechat_template"),
    "xcxdingyue": ("MA", "wechat_template"),
    "xuanxingbao": ("媒体渠道", "选型宝"),
    "d1net": ("传统媒体", "wechat-dsp"),
    "douyin-laike": ("付费媒体", "douyin-laike"),
    "toutiao": ("付费媒体", "toutiao"),
    "solutioncom": ("营销渠道", "wechat_dellemcsolution"),
    "solution_mb": ("营销渠道", "wechat_dellemcsolution"),
    "wechat_dellemcsolution": ("营销渠道", "wechat_dellemcsolution"),
    "wechat_dellemcsolution_menu": ("营销渠道", "wechat_dellemcsolution"),
    "wechat_dellemcsolution_tuiwen": ("营销渠道", "wechat_dellemcsolution"),
    "wechat_dellkeji": ("营销渠道", "wechat_dellkeji"),
    "wechat_dellkeji_menu": ("营销渠道", "wechat_dellkeji"),
    "wechat_dellkeji_tuiwen": ("营销渠道", "wechat_dellkeji"),
    "wechat_dellqicai": ("营销渠道", "wechat_dellqicai"),
    "wechat_dellqicai_menu": ("营销渠道", "wechat_dellqicai"),
    "wechat_dellqicai_tuiwen": ("营销渠道", "wechat_dellqicai"),
    "wechat_dellshangyong": ("营销渠道", "wechat_dellshangyong"),
    "wechat_dellshangyongsolution_500integral": ("营销渠道", "wechat_dellshangyong"),
    "wechat_shequn": ("营销渠道", "wechat_shequn"),
    "shequn": ("营销渠道", "wechat_shequn"),
    "dell-call": ("主动外呼", "主动外呼"),
}


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\xa0", " ").strip().lower()
    return re.sub(r"\s+", " ", text)


def split_tokens(raw: str) -> list[str]:
    cleaned = normalize_text(raw)
    cleaned = cleaned.replace("字样", "")
    parts = re.split(r"\s*/\s*|\s+或\s+|,", cleaned)
    return [part.strip() for part in parts if part.strip()]


@dataclass(frozen=True)
class RuleCondition:
    field: str
    operator: str
    tokens: tuple[str, ...]

    def matches(self, row: dict[str, Any]) -> bool:
        value = normalize_text(row.get(self.field))
        if self.operator == "contains":
            return any(token_matches(self.field, token, value) for token in self.tokens)
        if self.operator == "not_contains":
            return all(token not in value for token in self.tokens)
        if self.operator == "equals":
            return value in self.tokens
        raise ValueError(f"Unsupported operator: {self.operator}")


def token_matches(field: str, token: str, value: str) -> bool:
    if token not in value:
        return False
    if field == "utm_source":
        if value.endswith(("_xcx", "-xcx")) and token in {"wechat-dsp", "xiaohongshu"}:
            return False
        if value == "douyin-laike" and token == "douyin":
            return False
    return True


@dataclass(frozen=True)
class ChannelRule:
    row_number: int
    raw_rule: str
    channel_classification: str
    channel: str
    source: str
    conditions: tuple[RuleCondition, ...]

    def matches(self, row: dict[str, Any]) -> bool:
        return all(condition.matches(row) for condition in self.conditions)


def parse_rule_line(line: str) -> list[RuleCondition]:
    text = normalize_text(line)
    if not text:
        return []
    text = text.replace("其他参数任意", "").strip()
    text = text.replace("并且", "").strip()
    if not re.search(r"[a-zA-Z_]", text):
        return []

    match = re.search(r"([a-zA-Z_]+)\s*【(包含|不包含|等于|包括)】\s*=?\s*(.+)", text)
    if match:
        field_name, operator, raw_tokens = match.groups()
        field = FIELD_ALIASES.get(field_name.lower())
        if not field:
            raise ValueError(f"Unknown rule field: {field_name}")
        op_map = {
            "包含": "contains",
            "包括": "contains",
            "不包含": "not_contains",
            "等于": "equals",
        }
        extra_conditions: list[RuleCondition] = []
        cleaned_tokens = raw_tokens
        negative_match = re.search(r"(.+?)【不含(.+?)】", raw_tokens)
        if negative_match:
            cleaned_tokens = negative_match.group(1)
            extra_conditions.append(
                RuleCondition(field=field, operator="not_contains", tokens=tuple(split_tokens(negative_match.group(2))))
            )
        return [
            RuleCondition(field=field, operator=op_map[operator], tokens=tuple(split_tokens(cleaned_tokens))),
            *extra_conditions,
        ]

    match = re.search(r"([a-zA-Z_]+)\s*=\s*(.+)", text)
    if match:
        field_name, raw_tokens = match.groups()
        field = FIELD_ALIASES.get(field_name.lower())
        if not field:
            raise ValueError(f"Unknown equality field: {field_name}")
        return [RuleCondition(field=field, operator="equals", tokens=tuple(split_tokens(raw_tokens)))]

    return []


def parse_rule_block(raw_rule: str) -> tuple[RuleCondition, ...]:
    lines: list[str] = []
    for raw_line in str(raw_rule).splitlines():
        for part in re.split(r"[，,]", raw_line):
            stripped = part.strip()
            if stripped:
                lines.append(stripped)
    conditions: list[RuleCondition] = []
    for line in lines:
        conditions.extend(parse_rule_line(line))
    return tuple(conditions)


class ChannelRuleEngine:
    def __init__(self, rules: list[ChannelRule]) -> None:
        self.rules = rules

    @classmethod
    def from_workbook(cls, workbook_path: Path) -> "ChannelRuleEngine":
        wb = load_workbook(workbook_path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rules: list[ChannelRule] = []
        for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            raw_rule, classification, channel, source = row[:4]
            if not raw_rule:
                continue
            rules.append(
                ChannelRule(
                    row_number=row_number,
                    raw_rule=str(raw_rule),
                    channel_classification=str(classification or "").strip(),
                    channel=str(channel or "").strip(),
                    source=str(source or "").strip(),
                    conditions=parse_rule_block(str(raw_rule)),
                )
            )
        return cls(rules)

    def match(self, row: dict[str, Any]) -> tuple[str, str, int | None]:
        for rule in self.rules:
            if rule.matches(row):
                fallback = self._fallback_match(row)
                if fallback and rule.channel_classification == "自然流量":
                    return fallback[0], fallback[1], None
                return rule.channel_classification, rule.channel, rule.row_number
        fallback = self._fallback_match(row)
        if fallback:
            return fallback[0], fallback[1], None
        return "自然流量", "自然流量", None

    def _fallback_match(self, row: dict[str, Any]) -> tuple[str, str] | None:
        utm_source = normalize_text(row.get("utm_source"))
        if not utm_source:
            return None
        if "wechat_dellshangyongsolution" in utm_source:
            return ("营销渠道", "wechat_dellshangyong")
        if utm_source in SOURCE_FALLBACK_RULES:
            return SOURCE_FALLBACK_RULES[utm_source]
        return None
