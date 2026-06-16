from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


@dataclass(frozen=True)
class SourceRecord:
    source_name: str
    data: dict[str, Any]


def _load_sheet(path: Path, read_only: bool = True):
    return load_workbook(path, read_only=read_only, data_only=True)


def _open_worksheet(path: Path, sheet_name: str, header_row: int):
    wb = _load_sheet(path, read_only=True)
    ws = wb[sheet_name]
    if ws.max_row <= header_row and "finished-solution" in path.name:
        wb = _load_sheet(path, read_only=False)
        ws = wb[sheet_name]
    return ws


def _iter_worksheet_rows(path: Path, sheet_name: str, header_row: int, data_start_row: int, start_col: int = 1) -> Iterable[dict[str, Any]]:
    ws = _open_worksheet(path, sheet_name, header_row)

    header_values: list[Any] = []
    for row in ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True):
        header_values = list(row[start_col - 1 :])
    normalized_headers = [str(header).strip() if header is not None else "" for header in header_values]

    for values_row in ws.iter_rows(min_row=data_start_row, values_only=True):
        values = list(values_row[start_col - 1 : start_col - 1 + len(normalized_headers)])
        if not any(value not in (None, "") for value in values):
            continue
        row = {
            normalized_headers[idx]: values[idx]
            for idx in range(len(normalized_headers))
            if normalized_headers[idx]
        }
        yield row


def _detect_chat_layout(path: Path, sheet_name: str = "sheet1") -> tuple[int, int, int]:
    ws = _open_worksheet(path, sheet_name, header_row=2)
    candidates = [
        (1, 1, "MKT ID"),
        (2, 2, "MKT ID"),
    ]
    for header_row, start_col, marker in candidates:
        value = ws.cell(header_row, start_col).value
        if str(value or "").strip() == marker:
            return header_row, header_row + 1, start_col
    raise ValueError(f"Cannot detect chat layout for {path}")


def _detect_phone400_layout(path: Path, sheet_name: str = "sheet1") -> tuple[int, int, int]:
    ws = _open_worksheet(path, sheet_name, header_row=1)
    candidates = [
        (1, 1, "媒体来源"),
        (2, 1, "媒体来源"),
    ]
    for header_row, start_col, marker in candidates:
        value = ws.cell(header_row, start_col).value
        if str(value or "").strip() == marker:
            return header_row, header_row + 1, start_col
    raise ValueError(f"Cannot detect phone 400 layout for {path}")


def read_q2_website(path: Path, filtered_counts: dict[str, int] | None = None) -> Iterable[SourceRecord]:
    for row in _iter_worksheet_rows(path, "FY27Q2数据明细-唯一", header_row=1, data_start_row=2):
        yield SourceRecord(source_name="website_q2", data=row)


def read_q1_website(path: Path, filtered_counts: dict[str, int] | None = None) -> Iterable[SourceRecord]:
    for row in _iter_worksheet_rows(path, "FY27Q1数据明细-唯一", header_row=1, data_start_row=2):
        module_value = str(row.get("LV2 来源模块") or row.get("栏目模块") or row.get("栏目模块 ") or "")
        form_value = str(row.get("来源表单") or "")
        if "商城订单" in module_value or "商城订单" in form_value:
            if filtered_counts is not None:
                filtered_counts["website_q1_mall_order_filtered"] = filtered_counts.get("website_q1_mall_order_filtered", 0) + 1
            continue
        yield SourceRecord(source_name="website_q1", data=row)


def read_fy26q4_weekly_website(path: Path, filtered_counts: dict[str, int] | None = None) -> Iterable[SourceRecord]:
    for row in _iter_worksheet_rows(path, "FY26Q4数据明细-周报用唯一", header_row=1, data_start_row=2):
        module_value = str(row.get("LV2 来源模块") or row.get("栏目模块") or row.get("栏目模块 ") or "")
        form_value = str(row.get("来源表单") or "")
        if "商城订单" in module_value or "商城订单" in form_value:
            if filtered_counts is not None:
                filtered_counts["website_fy26q4_weekly_mall_order_filtered"] = (
                filtered_counts.get("website_fy26q4_weekly_mall_order_filtered", 0) + 1
                )
            continue
        yield SourceRecord(source_name="website_fy26q4_weekly", data=row)


def read_fy26q3_combined_website(path: Path, filtered_counts: dict[str, int] | None = None) -> Iterable[SourceRecord]:
    for sheet_name in ("数据明细", "腾讯数据"):
        for row in _iter_worksheet_rows(path, sheet_name, header_row=1, data_start_row=2):
            module_value = str(row.get("LV2 来源模块") or row.get("栏目模块") or row.get("栏目模块 ") or "")
            form_value = str(row.get("来源表单") or "")
            if "商城订单" in module_value or "商城订单" in form_value:
                if filtered_counts is not None:
                    key = f"website_fy26q3_{sheet_name}_mall_order_filtered"
                    filtered_counts[key] = filtered_counts.get(key, 0) + 1
                continue
            yield SourceRecord(source_name="website_fy26q3_combined", data=row)


def read_fy26q2_sino_website(path: Path, filtered_counts: dict[str, int] | None = None) -> Iterable[SourceRecord]:
    for row in _iter_worksheet_rows(path, "Sino外呼数据明细", header_row=1, data_start_row=2):
        module_value = str(row.get("LV2 来源模块") or row.get("栏目模块") or row.get("栏目模块 ") or "")
        form_value = str(row.get("来源表单") or "")
        if "商城订单" in module_value or "商城订单" in form_value:
            if filtered_counts is not None:
                key = "website_fy26q2_sino_mall_order_filtered"
                filtered_counts[key] = filtered_counts.get(key, 0) + 1
            continue
        yield SourceRecord(source_name="website_fy26q2_sino", data=row)


def read_chat(path: Path) -> Iterable[SourceRecord]:
    header_row, data_start_row, start_col = _detect_chat_layout(path, "sheet1")
    for row in _iter_worksheet_rows(path, "sheet1", header_row=header_row, data_start_row=data_start_row, start_col=start_col):
        yield SourceRecord(source_name="chat", data=row)


def read_phone_400(path: Path) -> Iterable[SourceRecord]:
    header_row, data_start_row, start_col = _detect_phone400_layout(path, "sheet1")
    for row in _iter_worksheet_rows(path, "sheet1", header_row=header_row, data_start_row=data_start_row, start_col=start_col):
        yield SourceRecord(source_name="phone_400", data=row)


def read_chat_phone_combined(path: Path) -> Iterable[SourceRecord]:
    for row in _iter_worksheet_rows(path, "Sheet1", header_row=1, data_start_row=2):
        yield SourceRecord(source_name="chat_phone_combined", data=row)


def read_orders(path: Path) -> Iterable[SourceRecord]:
    for row in _iter_worksheet_rows(path, "orders", header_row=1, data_start_row=2, start_col=1):
        yield SourceRecord(source_name="order", data=row)
