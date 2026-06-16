#!/usr/bin/env python3
import argparse
import csv
import json
import os
import sys
from collections import Counter
from datetime import date, datetime

from openpyxl import load_workbook

REMOVE_SOURCES = ["wechat-dsp-xcx", "wechat-dsp", "douyin", "kuaishou", "toutiao", "xiaohongshu", "xiaohongshu_xcx"]
KEEP_SOURCE = "wechat-dsp"
SOURCE_COLUMN_INDEX = 42
SOURCE_COLUMN_NAMES = {
    "媒体来源",
    "数据来源",
    "source",
}
NORMALIZED_REMOVE_SOURCES = {value.lower() for value in REMOVE_SOURCES}
NORMALIZED_KEEP_SOURCE = KEEP_SOURCE.lower()


def parse_args():
    parser = argparse.ArgumentParser(description="Split leads export into two output tables.")
    parser.add_argument("--input", required=True, help="Source CSV or XLSX path")
    parser.add_argument("--output-dir", required=True, help="Directory for generated CSV files")
    return parser.parse_args()


def extract_date_from_filename(filepath):
    filename = os.path.basename(filepath)
    if "_by_" in filename:
        suffix = filename.split("_by_")[-1]
        return os.path.splitext(suffix)[0]
    return datetime.now().strftime("%Y%m%d")


def read_csv_rows(filepath):
    with open(filepath, "r", encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.reader(handle))
    if not rows:
        raise ValueError("CSV 文件为空")
    return rows[0], rows[1:]


def serialize_cell_value(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    return value


def read_xlsx_rows(filepath):
    workbook = load_workbook(filepath, read_only=True, data_only=True)
    try:
        worksheet = workbook[workbook.sheetnames[0]]
        rows = [[serialize_cell_value(cell) for cell in row] for row in worksheet.iter_rows(values_only=True)]
    finally:
        workbook.close()

    if not rows:
        raise ValueError("Excel 文件为空")
    return rows[0], rows[1:]


def read_rows(filepath):
    ext = os.path.splitext(filepath)[1].lower()
    if ext == ".csv":
        return "csv", *read_csv_rows(filepath)
    if ext == ".xlsx":
        return "xlsx", *read_xlsx_rows(filepath)
    raise ValueError("当前插件仅支持 CSV 或 XLSX 文件")


def write_csv(filepath, header, rows):
    with open(filepath, "w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(header)
        writer.writerows(rows)


def normalize_text(value):
    return str(value or "").strip()


def normalize_source_value(value):
    return normalize_text(value).lower()


def normalize_header(value):
    return " ".join(normalize_text(value).replace("\r", "\n").split()).lower()


def get_row_value(row, index):
    if index < 0 or index >= len(row):
        return ""
    return normalize_text(row[index])


def detect_source_column_index(header, fallback_index):
    normalized_targets = {normalize_header(name) for name in SOURCE_COLUMN_NAMES}
    for index, value in enumerate(header):
        if normalize_header(value) in normalized_targets:
            return index
    if len(header) > fallback_index:
        return fallback_index
    raise ValueError("未找到来源列（媒体来源 / 数据来源 / source）")


def build_split_result(header, data_rows, input_type):
    fallback_index = 0 if input_type == "csv" else SOURCE_COLUMN_INDEX
    source_index = detect_source_column_index(header, fallback_index)
    source_column_label = normalize_text(header[source_index]) or f"Column {source_index + 1}"
    source_counter = Counter()
    for row in data_rows:
        source_counter[get_row_value(row, source_index)] += 1

    table1_rows = [row for row in data_rows if normalize_source_value(get_row_value(row, source_index)) not in NORMALIZED_REMOVE_SOURCES]
    table2_rows = [row for row in data_rows if normalize_source_value(get_row_value(row, source_index)) == NORMALIZED_KEEP_SOURCE]

    return {
        "header": header,
        "table1_rows": table1_rows,
        "table2_rows": table2_rows,
        "table1_description": "Removes wechat-dsp-xcx, wechat-dsp, douyin, kuaishou, toutiao, xiaohongshu, and xiaohongshu_xcx.",
        "table2_description": "Keeps only wechat-dsp for the downstream handoff.",
        "distribution": dict(source_counter),
        "distribution_label": f"Source Distribution · {source_column_label}",
        "input_type": input_type,
        "source_column_index": source_index,
        "source_column_label": source_column_label,
    }


def main():
    args = parse_args()
    os.makedirs(args.output_dir, exist_ok=True)

    input_type, header, data_rows = read_rows(args.input)
    if not header:
        raise ValueError("上传表表头为空")

    result = build_split_result(header, data_rows, input_type)

    date_str = extract_date_from_filename(args.input)
    table1_filename = f"leads_form_by_{date_str}#{len(result['table1_rows'])}.csv"
    table2_filename = f"leads_form_by_{date_str}#{len(result['table2_rows'])}-to ouyang.csv"

    write_csv(os.path.join(args.output_dir, table1_filename), result["header"], result["table1_rows"])
    write_csv(os.path.join(args.output_dir, table2_filename), result["header"], result["table2_rows"])

    payload = {
        "total": len(data_rows),
        "table1_count": len(result["table1_rows"]),
        "table2_count": len(result["table2_rows"]),
        "source_distribution": result["distribution"],
        "distribution_label": result["distribution_label"],
        "input_type": result["input_type"],
        "outputs": [
            {
                "key": "table1",
                "label": "Table 1",
                "description": result["table1_description"],
                "filename": table1_filename,
                "count": len(result["table1_rows"]),
            },
            {
                "key": "table2",
                "label": "Table 2",
                "description": result["table2_description"],
                "filename": table2_filename,
                "count": len(result["table2_rows"]),
            },
        ],
    }
    print(json.dumps(payload, ensure_ascii=False))


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(str(exc), file=sys.stderr)
        sys.exit(1)
