#!/usr/bin/env python3
import argparse
import json
import os
import re
from typing import Iterable, List

from openpyxl import Workbook, load_workbook


TARGET_COLUMNS = [
    "Mobile",
    "Email",
    "Company",
    "Lead : Account : Affinity Account ID",
    "State/Province",
    "City",
    "Created Date",
    "Additional Comments",
    "Lead Status",
    "Reason",
    "Lead Rating",
    "Allocaida ID",
    "call_type",
]

DEFAULT_SHEET_CANDIDATES = ("Sheet1", "明细")


def parse_args():
    parser = argparse.ArgumentParser(description="Build a target-style Signal Leads workbook from SFDC export.")
    parser.add_argument("--input", required=True, help="Source workbook path")
    parser.add_argument("--output", required=True, help="Output workbook path")
    parser.add_argument("--sheet", default="", help="Source sheet name. Defaults to Sheet1, then 明细.")
    return parser.parse_args()


def normalize_text(value):
    if value is None:
        return ""
    return str(value).strip()


def clean_mobile(value):
    text = re.sub(r"\D", "", normalize_text(value))
    if text.startswith("86"):
        text = text[2:]
    if len(text) != 11 or not text.startswith("1"):
        return ""
    return text


def clean_state(value):
    text = normalize_text(value)
    for suffix in ("省", "市"):
        if text.endswith(suffix):
            return text[:-1]
    return text


def clean_city(value):
    text = normalize_text(value)
    if text.endswith("市"):
        return text[:-1]
    return text


def extract_allocaida_id(value):
    text = normalize_text(value)
    compact_text = re.sub(r"\s+", "", text)
    if "eleads" in compact_text and "备注没ID" in compact_text:
        return "1"
    if "邮箱" in compact_text and "其他Campaign" in compact_text and "重复" in compact_text:
        return "2"
    if compact_text in {"3445890", "3445891", "3445892"}:
        return compact_text
    match = re.search(r"(\d+)$", text)
    if text.startswith("eleads-") and match:
        return match.group(1)
    return ""


def normalize_call_type(value):
    text = normalize_text(value)
    if text in {"Signal", "Signal(待确定)"}:
        return "Signals"
    return text


def load_rows(input_path: str, sheet_name: str):
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    try:
        selected_sheet = sheet_name
        if selected_sheet:
            if selected_sheet not in workbook.sheetnames:
                raise ValueError(f"Sheet not found: {selected_sheet}")
        else:
            selected_sheet = next((name for name in DEFAULT_SHEET_CANDIDATES if name in workbook.sheetnames), "")
            if not selected_sheet:
                raise ValueError(f"Sheet not found: {' / '.join(DEFAULT_SHEET_CANDIDATES)}")
        worksheet = workbook[selected_sheet]
        rows = list(worksheet.iter_rows(values_only=True))
    finally:
        workbook.close()

    if not rows:
        raise ValueError("Source sheet is empty")
    return selected_sheet, rows[0], rows[1:]


def build_header_index(header: Iterable[str]):
    return {normalize_text(value): index for index, value in enumerate(header)}


def get_cell(row: List[object], index: int):
    if index < 0 or index >= len(row):
        return ""
    value = row[index]
    return "" if value is None else value


def build_output_rows(source_rows, header_index):
    output_rows = []
    for row in source_rows:
        output_row = []
        for column in TARGET_COLUMNS:
            if column == "Allocaida ID":
                output_row.append(extract_allocaida_id(get_cell(row, header_index["是否eleads"])))
                continue
            if column == "call_type":
                output_row.append(normalize_call_type(get_cell(row, header_index["Column1"])) if "Column1" in header_index else "")
                continue
            value = get_cell(row, header_index[column])
            if column == "Mobile":
                output_row.append(clean_mobile(value))
                continue
            if column == "State/Province":
                output_row.append(clean_state(value))
                continue
            if column == "City":
                output_row.append(clean_city(value))
                continue
            output_row.append(value)
        output_rows.append(output_row)
    return output_rows


def write_workbook(output_path: str, rows: List[List[object]]):
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    worksheet.append(TARGET_COLUMNS)
    for row in rows:
        worksheet.append(row)
    workbook.save(output_path)


def main():
    args = parse_args()
    selected_sheet, header, source_rows = load_rows(args.input, args.sheet)
    header_index = build_header_index(header)

    missing_columns = [
        column for column in TARGET_COLUMNS if column not in {"Allocaida ID", "call_type"} and column not in header_index
    ]
    if "是否eleads" not in header_index:
        missing_columns.append("是否eleads")
    if missing_columns:
        raise ValueError(f"Missing required columns: {', '.join(sorted(set(missing_columns)))}")

    output_rows = build_output_rows(source_rows, header_index)
    write_workbook(args.output, output_rows)
    payload = {
        "sheet": selected_sheet,
        "source_total_rows": len(source_rows),
        "output_rows": len(output_rows),
        "output_path": args.output,
        "output_filename": os.path.basename(args.output),
        "non_empty_mobile_count": sum(1 for row in output_rows if normalize_text(row[0])),
        "allocaida_non_empty_count": sum(1 for row in output_rows if normalize_text(row[11])),
        "call_type_non_empty_count": sum(1 for row in output_rows if normalize_text(row[12])),
    }
    print(json.dumps(payload, ensure_ascii=False))


if __name__ == "__main__":
    main()
