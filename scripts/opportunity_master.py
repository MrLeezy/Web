#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import re
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from opportunity_master_lib import (
    BuildStats,
    ChannelRuleEngine,
    apply_sfdc_call_type_override,
    dedupe_rows,
    export_workbook,
    normalize_email_for_match,
    normalize_phone_for_match,
    read_chat,
    read_chat_phone_combined,
    read_fy26q2_sino_website,
    read_fy26q3_combined_website,
    read_fy26q4_weekly_website,
    read_orders,
    read_phone_400,
    read_q1_website,
    read_q2_website,
    should_drop_row,
    transform_record,
)


SOURCE_READERS = {
    "website_q2": read_q2_website,
    "website_q1": read_q1_website,
    "website_fy26q4_weekly": read_fy26q4_weekly_website,
    "website_fy26q3_combined": read_fy26q3_combined_website,
    "website_fy26q2_sino": read_fy26q2_sino_website,
    "chat": read_chat,
    "chat_phone_combined": read_chat_phone_combined,
    "phone_400": read_phone_400,
    "order": read_orders,
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build opportunity master workbook from uploaded source files.")
    parser.add_argument("--mapping-table", required=True, help="Path to the mapping workbook")
    parser.add_argument("--channel-mapping", required=True, help="Path to the channel mapping workbook")
    parser.add_argument("--output-dir", required=True, help="Directory for generated workbook")
    parser.add_argument("--run-label", default="", help="Optional label appended to the output filename")
    parser.add_argument("--website-q2-file", action="append", default=[], help="Website FY27Q2 source file")
    parser.add_argument("--website-q1-file", action="append", default=[], help="Website FY27Q1 source file")
    parser.add_argument("--website-fy26q4-weekly-file", action="append", default=[], help="Website FY26Q4 weekly source file")
    parser.add_argument("--website-fy26q3-combined-file", action="append", default=[], help="Website FY26Q3 combined source file")
    parser.add_argument("--website-fy26q2-sino-file", action="append", default=[], help="Website FY26Q2 Sino source file")
    parser.add_argument("--chat-file", action="append", default=[], help="Chat source file")
    parser.add_argument("--chat-phone-combined-file", action="append", default=[], help="Combined chat and 400 source file")
    parser.add_argument("--phone-400-file", action="append", default=[], help="Phone 400 source file")
    parser.add_argument("--order-file", action="append", default=[], help="Order source file")
    parser.add_argument("--sfdc-signal-file", action="append", default=[], help="SFDC signal mapping workbook for call type overrides")
    return parser.parse_args()


def load_target_fields(mapping_path: Path) -> list[str]:
    wb = load_workbook(mapping_path, read_only=True, data_only=True)
    ws = wb["商机映射规则"] if "商机映射规则" in wb.sheetnames else wb[wb.sheetnames[0]]
    fields: list[str] = []
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0]:
            fields.append(str(row[0]).strip())
    if "Campaign Name" not in fields:
        fields.append("Campaign Name")
    if "Allocaida ID" not in fields:
        fields.append("Allocaida ID")
    wb.close()
    return fields


def clean_label(raw: str) -> str:
    text = re.sub(r"[^0-9A-Za-z\u4e00-\u9fff_-]+", "_", raw.strip())
    text = re.sub(r"_+", "_", text).strip("_")
    return text[:80] or "manual"


def build_output_path(output_dir: Path, run_label: str) -> Path:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"商机数据整合结果_{timestamp}.xlsx"
    if run_label:
        filename = f"商机数据整合结果_{clean_label(run_label)}_{timestamp}.xlsx"
    return output_dir / timestamp / filename


def list_source_plan(args: argparse.Namespace) -> list[tuple[str, Path]]:
    source_files = [
        ("website_q2", args.website_q2_file),
        ("website_q1", args.website_q1_file),
        ("website_fy26q4_weekly", args.website_fy26q4_weekly_file),
        ("website_fy26q3_combined", args.website_fy26q3_combined_file),
        ("website_fy26q2_sino", args.website_fy26q2_sino_file),
        ("chat", args.chat_file),
        ("chat_phone_combined", args.chat_phone_combined_file),
        ("phone_400", args.phone_400_file),
        ("order", args.order_file),
    ]
    plan: list[tuple[str, Path]] = []
    for source_name, paths in source_files:
        for raw_path in paths:
            plan.append((source_name, Path(raw_path)))
    return plan


def list_sfdc_signal_files(args: argparse.Namespace) -> list[Path]:
    return [Path(raw_path) for raw_path in args.sfdc_signal_file]


def normalized_header_index(header_row: tuple[object, ...]) -> dict[str, int]:
    return {
        re.sub(r"\s+", "", str(value or "")).strip().lower(): index
        for index, value in enumerate(header_row)
    }


def cell_value(row: tuple[object, ...], index: int) -> object:
    if index < 0 or index >= len(row):
        return ""
    return row[index]


def load_sfdc_signal_match_keys(paths: list[Path], stats: BuildStats) -> tuple[set[str], set[str]]:
    phone_matches: set[str] = set()
    email_matches: set[str] = set()

    for workbook_path in paths:
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        try:
            sheet_name = "Sheet1" if "Sheet1" in workbook.sheetnames else workbook.sheetnames[0]
            worksheet = workbook[sheet_name]
            row_iter = worksheet.iter_rows(values_only=True)
            header_row = next(row_iter, None)
            if header_row is None:
                continue

            header_index = normalized_header_index(header_row)
            missing_columns = [name for name in ("mobile", "email") if name not in header_index]
            if missing_columns:
                raise ValueError(f"SFDC signal 文件缺少必要列: {', '.join(missing_columns)}")

            for row in row_iter:
                phone_key = normalize_phone_for_match(cell_value(row, header_index["mobile"]))
                email_key = normalize_email_for_match(cell_value(row, header_index["email"]))
                if phone_key:
                    phone_matches.add(phone_key)
                if email_key:
                    email_matches.add(email_key)
                if phone_key or email_key:
                    stats.filtered_counts["sfdc_signal_lookup_rows"] += 1
        finally:
            workbook.close()

    return phone_matches, email_matches


def build_payload(output_path: Path, rows: list[dict[str, str]], stats: BuildStats, source_files: list[tuple[str, Path]]) -> dict[str, object]:
    return {
        "outputFile": str(output_path),
        "outputFilename": output_path.name,
        "totalRows": len(rows),
        "sourceCounts": dict(stats.output_counts),
        "inputCounts": dict(stats.input_counts),
        "invalidCounts": dict(stats.invalid_counts),
        "filteredCounts": dict(stats.filtered_counts),
        "channelCounts": dict(stats.channel_counts),
        "sourceFiles": [
            {
                "sourceName": source_name,
                "path": str(file_path),
                "filename": file_path.name,
            }
            for source_name, file_path in source_files
        ],
    }


def main() -> None:
    args = parse_args()
    mapping_path = Path(args.mapping_table)
    channel_mapping_path = Path(args.channel_mapping)
    output_dir = Path(args.output_dir)

    source_plan = list_source_plan(args)
    sfdc_signal_files = list_sfdc_signal_files(args)
    if not source_plan:
        raise ValueError("至少需要一份已识别的来源文件")

    target_fields = load_target_fields(mapping_path)
    channel_engine = ChannelRuleEngine.from_workbook(channel_mapping_path)
    stats = BuildStats()
    phone_matches, email_matches = load_sfdc_signal_match_keys(sfdc_signal_files, stats)
    rows: list[dict[str, str]] = []

    for source_name, file_path in source_plan:
        reader = SOURCE_READERS[source_name]
        iterable = reader(file_path, stats.filtered_counts) if source_name.startswith("website_") else reader(file_path)
        for record in iterable:
            stats.input_counts[source_name] += 1
            transformed = transform_record(record.source_name, record.data, target_fields, stats, channel_engine)
            if should_drop_row(transformed):
                stats.filtered_counts["order_empty_rows_dropped"] += 1
                continue
            if phone_matches or email_matches:
                transformed = apply_sfdc_call_type_override(transformed, phone_matches, email_matches, stats)
            rows.append(transformed)
            stats.output_counts[source_name] += 1

    before_dedupe = len(rows)
    rows = dedupe_rows(rows)
    stats.filtered_counts["leads_id_deduped_rows"] += before_dedupe - len(rows)

    output_path = build_output_path(output_dir, args.run_label)
    export_workbook(output_path, target_fields, rows, stats)
    print(json.dumps(build_payload(output_path, rows, stats, source_plan), ensure_ascii=False))


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(str(exc), file=sys.stderr)
        sys.exit(1)
